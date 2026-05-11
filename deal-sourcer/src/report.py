"""
src/report.py
Excel report generator -- toggle-driven formula build.

4 tabs (in order):
  1. Toggles  -- editable OCC_TOGGLE / OPEX_TOGGLE / CAP_TOGGLE; fixed assumptions
  2. PASS Deals -- historical valuation via Excel formulas driven by Toggles
  3. FAIL Deals -- same structure + Fail Reason column
  4. EXCLUDED   -- simple listing, no formulas

Named ranges:
  OCC_TOGGLE  = Toggles!$B$4
  OPEX_TOGGLE = Toggles!$B$5
  CAP_TOGGLE  = Toggles!$B$6
  ACQ_MONTH   = Toggles!$B$8
  ACQ_DAY     = Toggles!$C$8
  ACQ_YEAR    = Toggles!$D$8

Formula columns reference named ranges so changing B4/B5/B6 on Toggles tab
recalculates all historical NOI, implied value, loan, DS, equity, and IRR
across both deal tabs.
"""

import sys
import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import config.assumptions as A

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_ORANGE = PatternFill("solid", fgColor="FFD28B")
_RED    = PatternFill("solid", fgColor="FFC7CE")
_DKBLUE = PatternFill("solid", fgColor="1F3864")
_LTBLUE = PatternFill("solid", fgColor="DCE6F1")
_TOGGLE = PatternFill("solid", fgColor="FFFF00")   # editable cells

_HFONT  = Font(bold=True, color="FFFFFF", size=10)
_BOLD   = Font(bold=True)
_THIN   = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
_CTR  = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_RENT_FILL = {
    "actual":                   None,
    "estimated_market_vintage": _YELLOW,
    "estimated_market_avg":     _ORANGE,
    "no_rent_data":             _RED,
}

# ---------------------------------------------------------------------------
# Pre-computed column letter lookup (1-60)
# ---------------------------------------------------------------------------
_CL = {i: get_column_letter(i) for i in range(1, 61)}

# ---------------------------------------------------------------------------
# Column index constants for deal tabs
# ---------------------------------------------------------------------------
# ---- Static value columns (written from Python) ----
CA  = 1   # Property Name
CB  = 2   # Market
CC  = 3   # Submarket
CD  = 4   # State
CE  = 5   # Asset Type
CF  = 6   # Units
CG  = 7   # Vintage
CH  = 8   # Bldg Class
CI  = 9   # Owner
CJ  = 10  # Property Manager
CK  = 11  # Loan Maturity
CL  = 12  # Debt Type
CM  = 13  # Rent/Unit/Mo
CN  = 14  # Rent Method
CO  = 15  # Rent Comps

# ---- Toggle-driven formula columns ----
CP  = 16  # P  -- Hist GSR
CQ  = 17  # Q  -- Hist EGR (after occ toggle)
CR  = 18  # R  -- Hist Other Income
CS  = 19  # S  -- Hist RUBS
CT  = 20  # T  -- Hist Total Revenue
CU  = 21  # U  -- Hist Expenses (opex toggle)
CV  = 22  # V  -- Hist NOI
CW  = 23  # W  -- Hist Implied Value (Max Bid)
CX  = 24  # X  -- Max Bid Per Unit
CY  = 25  # Y  -- Loan Amount (65% of W)
CZ  = 26  # Z  -- Debt Svc Yr1 IO (Y * loan_rate)
CAA = 27  # AA -- Property Taxes Yr1 (W * 2%)

# ---- Python value columns (forward UW) ----
CAB = 28  # AB -- Mgmt Fee Yr1
CAC = 29  # AC -- Total Fwd Expenses Yr1
CAD = 30  # AD -- Fwd NOI Yr1
CAE = 31  # AE -- Fwd NOI Yr2
CAF = 32  # AF -- Fwd NOI Yr3
CAG = 33  # AG -- Fwd NOI Yr4
CAH = 34  # AH -- Fwd NOI Yr5

# ---- More formula columns ----
CAI = 35  # AI -- DS Yr1 formula (= Y * loan_rate, same as Z)
CAJ = 36  # AJ -- DS Yr4 amortizing formula
CAK = 37  # AK -- DS Yr5 amortizing formula
CAL = 38  # AL -- DSCR Yr1
CAM = 39  # AM -- DSCR Yr3
CAN = 40  # AN -- Exit Value (Fwd NOI Yr5 / exit_cap)
CAO = 41  # AO -- Loan Balance Month 60 (formula)
CAP = 42  # AP -- Net Sale Proceeds
CAQ = 43  # AQ -- Total Equity (W * 35%)
CAR = 44  # AR -- Deal IRR (XIRR)
CAS = 45  # AS -- Deal IRR % (= AR, formatted)
CAT = 46  # AT -- PASS/FAIL
CAU = 47  # AU -- Expense Ratio Yr1
CAV = 48  # AV -- Debt Yield Yr3
CAW = 49  # AW -- Going-In Cap
CAX = 50  # AX -- Loan Maturity Flag

# FAIL tab only
CAY = 51  # AY -- Fail Reason

# ---------------------------------------------------------------------------
# Header definitions: (col_index, label, width)
# ---------------------------------------------------------------------------
_HEADERS = [
    (CA,  "Property Name",        30),
    (CB,  "Market",               24),
    (CC,  "Submarket",            18),
    (CD,  "State",                 8),
    (CE,  "Asset Type",           12),
    (CF,  "Units",                 8),
    (CG,  "Vintage",               8),
    (CH,  "Bldg Class",           10),
    (CI,  "Owner",                24),
    (CJ,  "Property Manager",     22),
    (CK,  "Loan Maturity",        14),
    (CL,  "Debt Type",            14),
    (CM,  "Rent/Unit/Mo",         13),
    (CN,  "Rent Method",          22),
    (CO,  "Rent Comps",           10),
    # Toggle-driven
    (CP,  "Hist GSR",             14),
    (CQ,  "Hist EGR",             14),
    (CR,  "Hist Other Inc",       15),
    (CS,  "Hist RUBS",            13),
    (CT,  "Hist Tot Rev",         14),
    (CU,  "Hist Expenses",        14),
    (CV,  "Hist NOI",             14),
    (CW,  "Imp Value (Max Bid)",  20),
    (CX,  "Bid/Unit",             12),
    (CY,  "Loan Amount",          14),
    (CZ,  "DS Yr1 IO",            13),
    (CAA, "Prop Tax Yr1",         14),
    # Python values
    (CAB, "Mgmt Fee Yr1",         14),
    (CAC, "Fwd Exp Yr1",          14),
    (CAD, "NOI Yr1",              13),
    (CAE, "NOI Yr2",              13),
    (CAF, "NOI Yr3",              13),
    (CAG, "NOI Yr4",              13),
    (CAH, "NOI Yr5",              13),
    # Formula metrics
    (CAI, "DS Yr1",               13),
    (CAJ, "DS Yr4",               13),
    (CAK, "DS Yr5",               13),
    (CAL, "DSCR Yr1",             10),
    (CAM, "DSCR Yr3",             10),
    (CAN, "Exit Value",           14),
    (CAO, "Loan Bal Mo60",        14),
    (CAP, "Net Sale Proc",        14),
    (CAQ, "Total Equity",         14),
    (CAR, "Deal IRR",             12),
    (CAS, "Deal IRR %",           10),
    (CAT, "PASS/FAIL",            10),
    (CAU, "Exp Ratio",            10),
    (CAV, "Debt Yield Yr3",       14),
    (CAW, "Going-In Cap",         12),
    (CAX, "Mat Flag",             10),
]

_FAIL_EXTRA_HEADER = (CAY, "Fail Reason", 50)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _hdr_cell(ws, row, col, label, width=None):
    """Write a blue header cell."""
    c = ws.cell(row=row, column=col, value=label)
    c.fill = _DKBLUE
    c.font = _HFONT
    c.alignment = _CTR
    c.border = _THIN
    if width:
        ws.column_dimensions[_CL[col]].width = width
    return c


def _vc(ws, r, c, val, fmt=None, fill=None, left=False):
    """Write a value cell."""
    if val is None or (isinstance(val, float) and str(val) == "nan"):
        val = "--"
    cell = ws.cell(row=r, column=c, value=val)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    cell.border = _THIN
    cell.alignment = _LEFT if left else _CTR
    return cell


def _fc(ws, r, c, formula, fmt=None, fill=None):
    """Write a formula cell."""
    cell = ws.cell(row=r, column=c, value=formula)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    cell.border = _THIN
    cell.alignment = _CTR
    return cell


# ---------------------------------------------------------------------------
# SHEET 1 -- Toggles
# ---------------------------------------------------------------------------

def _write_toggles_tab(ws, meta: dict):
    """Build Toggles sheet with editable cells and fixed-assumption snapshot."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    # Row 1 -- title banner
    t = ws.cell(row=1, column=1, value="PORTICO ACQUISITION SCREENER")
    t.fill = _DKBLUE
    t.font = Font(bold=True, color="FFFFFF", size=14)
    t.alignment = _CTR
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30

    # Row 3 -- section header
    h = ws.cell(row=3, column=1, value="SCREENING TOGGLES")
    h.font = Font(bold=True, size=11)

    # Rows 4-6 -- editable toggle cells
    toggles = [
        (4, "Economic Occupancy", A.HIST_ECONOMIC_OCCUPANCY, "80% -- 95%"),
        (5, "OPEX Ratio",         A.HIST_OPEX_TOGGLE,        "40% -- 70%"),
        (6, "Going-In Cap Rate",  A.HIST_CAP_RATE_CONV,      "4% -- 7%"),
    ]
    for row, label, val, note in toggles:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _BOLD
        tc = ws.cell(row=row, column=2, value=val)
        tc.number_format = "0.00%"
        tc.fill = _TOGGLE
        tc.font = _BOLD
        tc.border = _THIN
        tc.alignment = _CTR
        ws.cell(row=row, column=3, value=note)

    # Row 7 -- blank spacer

    # Row 8 -- Acquisition Date inputs
    ws.cell(row=8, column=1, value="Acquisition Date").font = _BOLD
    b8 = ws.cell(row=8, column=2, value=6)
    b8.fill = _TOGGLE; b8.font = _BOLD; b8.border = _THIN; b8.alignment = _CTR
    c8 = ws.cell(row=8, column=3, value=1)
    c8.fill = _TOGGLE; c8.font = _BOLD; c8.border = _THIN; c8.alignment = _CTR
    d8 = ws.cell(row=8, column=4, value=2026)
    d8.fill = _TOGGLE; d8.font = _BOLD; d8.border = _THIN; d8.alignment = _CTR
    e8 = ws.cell(row=8, column=5, value="=DATE(D8,B8,C8)")
    e8.number_format = "MM/DD/YYYY"; e8.border = _THIN; e8.alignment = _CTR
    # Sub-labels for Month / Day / Year
    ws.cell(row=9, column=2, value="Month").alignment = _CTR
    ws.cell(row=9, column=3, value="Day").alignment = _CTR
    ws.cell(row=9, column=4, value="Year").alignment = _CTR

    # Row 10 -- Fixed Assumptions section
    ws.cell(row=10, column=1, value="FIXED ASSUMPTIONS").font = Font(bold=True, size=11)

    lr   = meta.get("loan_rate",     A.FALLBACK_LOAN_RATE)
    tsy  = meta.get("treasury_rate", A.FALLBACK_TREASURY)

    fixed = [
        (11, "Rent Growth",          A.RENT_GROWTH,            "0.00%"),
        (12, "Exit Cap (Conv)",       A.EXIT_CAP_CONV,          "0.00%"),
        (13, "Exit Cap (BTR TX)",     A.EXIT_CAP_BTR_TX_MAJOR,  "0.00%"),
        (14, "Exit Cap (BTR Other)",  A.EXIT_CAP_BTR_OTHER,     "0.00%"),
        (15, "LTV",                   A.LTV,                    "0.00%"),
        (16, "Loan Rate",             lr,                       "0.0000%"),
        (17, "5yr Treasury",          tsy,                      "0.0000%"),
        (18, "Hold Period",           "5 years",                None),
        (19, "IRR Threshold",         A.DEAL_IRR_MIN,           "0.00%"),
        (20, "Bad Debt Yr1",          A.BAD_DEBT_YEAR1,         "0.00%"),
        (21, "Bad Debt Yr2+",         A.BAD_DEBT_STABILIZED,    "0.00%"),
        (22, "Concessions Yr1",       A.CONCESSIONS_YEAR1,      "0.00%"),
        (23, "Vacancy",               A.VACANCY_RATE,           "0.00%"),
    ]
    for row, label, val, fmt in fixed:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        if fmt:
            c.number_format = fmt
        c.fill = _LTBLUE

    # Row 25 -- Per Door section
    ws.cell(row=25, column=1, value="PER DOOR ASSUMPTIONS").font = Font(bold=True, size=11)

    per_door = [
        (26, "Admin",             A.ADMIN_PER_DOOR,              "$#,##0"),
        (27, "Marketing",         A.MARKETING_PER_DOOR,          "$#,##0"),
        (28, "Payroll",           A.PAYROLL_PER_DOOR,            "$#,##0"),
        (29, "R&M",               A.RM_PER_DOOR,                 "$#,##0"),
        (30, "Contract Services", A.CONTRACT_SERVICES_PER_DOOR,  "$#,##0"),
        (31, "CapEx Reserve",     A.CAPEX_PER_DOOR,              "$#,##0"),
        (32, "Insurance",         A.INSURANCE_PER_UNIT_YEAR,     "$#,##0"),
        (33, "Utilities",         A.UTILITIES_PER_UNIT_YEAR,     "$#,##0"),
        (34, "RUBS (annual)",     A.RUBS_PER_DOOR_ANNUAL,        "$#,##0"),
        (35, "Other Income/mo",   A.OTHER_INCOME_PER_DOOR_MONTHLY, "$#,##0"),
        (36, "Mgmt Fee",          A.MGMT_FEE_PCT,                "0.00%"),
        (37, "Property Tax Rate", A.TAX_RATE_PCT,                "0.00%"),
    ]
    for row, label, val, fmt in per_door:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        c.fill = _LTBLUE

    # Row 39 -- Run Info section
    ws.cell(row=39, column=1, value="RUN INFO").font = Font(bold=True, size=11)

    run_info = [
        (40, "Run Date",       meta.get("run_ts", datetime.now().strftime("%Y-%m-%d %H:%M"))),
        (41, "Total Deals",    meta.get("total_deals", "--")),
        (42, "PASS Deals",     meta.get("n_pass",      "--")),
        (43, "FAIL Deals",     meta.get("n_fail",      "--")),
        (44, "EXCLUDED Deals", meta.get("n_excl",      "--")),
    ]
    for row, label, val in run_info:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Named ranges
# ---------------------------------------------------------------------------

def _add_named_ranges(wb):
    """Register OCC_TOGGLE, OPEX_TOGGLE, CAP_TOGGLE, ACQ_MONTH/DAY/YEAR as workbook-level named ranges."""
    ranges = {
        "OCC_TOGGLE":  "Toggles!$B$4",
        "OPEX_TOGGLE": "Toggles!$B$5",
        "CAP_TOGGLE":  "Toggles!$B$6",
        "ACQ_MONTH":   "Toggles!$B$8",
        "ACQ_DAY":     "Toggles!$C$8",
        "ACQ_YEAR":    "Toggles!$D$8",
    }
    for name, ref in ranges.items():
        dn = DefinedName(name=name, attr_text=ref)
        try:
            wb.defined_names.add(dn)
        except (AttributeError, TypeError, Exception):
            try:
                wb.defined_names[name] = dn
            except Exception:
                logger.warning("Could not register named range %s", name)


# ---------------------------------------------------------------------------
# SHEET 2/3 -- PASS / FAIL deal tabs
# ---------------------------------------------------------------------------

def _safe_num(v, default=0):
    """Return a clean float, or default if missing/NaN."""
    if v is None:
        return default
    try:
        f = float(v)
        return default if (f != f) else f   # NaN check
    except (TypeError, ValueError):
        return default


def _write_deal_row(ws, r: int, row_data: dict, is_fail: bool = False):
    """Write all columns for a single deal row (PASS or FAIL tab)."""

    # ---- pull values we need in formulas ----
    exit_cap = _safe_num(row_data.get("exit_cap"), A.EXIT_CAP_CONV)

    # ---- column letter shortcuts for this row ----
    def fl(col_idx): return f"{_CL[col_idx]}{r}"

    # ============================================================
    # Columns A–O: static values
    # ============================================================
    statics = [
        (CA,  row_data.get("property_name",   ""),  None,       None,    True),
        (CB,  row_data.get("market",          ""),  None,       None,    False),
        (CC,  row_data.get("submarket",       ""),  None,       None,    False),
        (CD,  row_data.get("state",           ""),  None,       None,    False),
        (CE,  row_data.get("asset_type",      ""),  None,       None,    False),
        (CF,  row_data.get("units"),               None,       None,    False),
        (CG,  row_data.get("vintage"),             None,       None,    False),
        (CH,  row_data.get("building_class",  ""),  None,       None,    False),
        (CI,  row_data.get("owner_name",      ""),  None,       None,    True),
        (CJ,  row_data.get("property_manager",""),  None,       None,    False),
        (CK,  row_data.get("loan_maturity_date","--"), None,    None,    False),
        (CL,  row_data.get("debt_type",       ""),  None,       None,    False),
        (CM,  _safe_num(row_data.get("monthly_rent_used")),
                                                    "$#,##0",  None,    False),
    ]
    for col, val, fmt, fill, left_align in statics:
        _vc(ws, r, col, val, fmt=fmt, fill=fill, left=left_align)

    # Rent Method -- colour coded
    rent_method_val = str(row_data.get("rent_method", "actual") or "actual").strip()
    rm_fill = _RENT_FILL.get(rent_method_val)
    _vc(ws, r, CN, rent_method_val, fill=rm_fill)

    # Rent Comps
    _vc(ws, r, CO, row_data.get("rent_comp_count", 0))

    # ============================================================
    # Columns P–AA: toggle-driven formula columns
    # ============================================================

    # P: Historical GSR  = Units * Rent * 12
    _fc(ws, r, CP,  f"={fl(CF)}*{fl(CM)}*12",             "$#,##0")

    # Q: Historical EGR  = GSR * OCC_TOGGLE
    _fc(ws, r, CQ,  f"={fl(CP)}*OCC_TOGGLE",              "$#,##0")

    # R: Historical Other Income = Units * $150 * 12
    _fc(ws, r, CR,  f"={fl(CF)}*150*12",                   "$#,##0")

    # S: Historical RUBS = Units * $840
    _fc(ws, r, CS,  f"={fl(CF)}*840",                      "$#,##0")

    # T: Total Historical Revenue = EGR + OI + RUBS
    _fc(ws, r, CT,  f"={fl(CQ)}+{fl(CR)}+{fl(CS)}",       "$#,##0")

    # U: Historical Expenses = Total Rev * OPEX_TOGGLE
    _fc(ws, r, CU,  f"={fl(CT)}*OPEX_TOGGLE",             "$#,##0")

    # V: Historical NOI = Rev - Exp
    _fc(ws, r, CV,  f"={fl(CT)}-{fl(CU)}",                "$#,##0")

    # W: Implied Value (Max Bid) = NOI / CAP_TOGGLE
    _fc(ws, r, CW,  f"=IFERROR({fl(CV)}/CAP_TOGGLE,0)",   "$#,##0")

    # X: Max Bid Per Unit = W / Units
    _fc(ws, r, CX,  f"=IFERROR({fl(CW)}/{fl(CF)},0)",     "$#,##0")

    # Y: Loan Amount = W * 65%
    _fc(ws, r, CY,  f"={fl(CW)}*0.65",                    "$#,##0")

    # Z: Debt Service Yr1 IO = Y * loan_rate (Toggles!$B$16)
    _fc(ws, r, CZ,  f"={fl(CY)}*Toggles!$B$16",           "$#,##0")

    # AA: Property Taxes Yr1 = W * 2%
    _fc(ws, r, CAA, f"={fl(CW)}*0.02",                    "$#,##0")

    # ============================================================
    # Columns AB–AH: Python values (forward UW output)
    # ============================================================
    py_vals = [
        (CAB, _safe_num(row_data.get("mgmt_fee_yr1")),      "$#,##0"),
        (CAC, _safe_num(row_data.get("total_expenses_yr1")),"$#,##0"),
        (CAD, _safe_num(row_data.get("noi_yr1")),           "$#,##0"),
        (CAE, _safe_num(row_data.get("noi_yr2")),           "$#,##0"),
        (CAF, _safe_num(row_data.get("noi_yr3")),           "$#,##0"),
        (CAG, _safe_num(row_data.get("noi_yr4")),           "$#,##0"),
        (CAH, _safe_num(row_data.get("noi_yr5")),           "$#,##0"),
    ]
    for col, val, fmt in py_vals:
        c = ws.cell(row=r, column=col, value=val)
        c.number_format = fmt
        c.border = _THIN
        c.alignment = _CTR

    # ============================================================
    # Columns AI–AX: formula-driven metrics
    # ============================================================

    # AI: DS Yr1 IO (formula, same as Z -- used in DSCR/XIRR refs)
    _fc(ws, r, CAI, f"={fl(CY)}*Toggles!$B$16",           "$#,##0")

    # AJ: DS Yr4 amortizing  = -PMT(rate/12, 360, loan) * 12
    _fc(ws, r, CAJ,
        f"=IFERROR(-PMT(Toggles!$B$16/12,360,{fl(CY)})*12,0)",
        "$#,##0")

    # AK: DS Yr5 amortizing (same payment -- level-pay)
    _fc(ws, r, CAK,
        f"=IFERROR(-PMT(Toggles!$B$16/12,360,{fl(CY)})*12,0)",
        "$#,##0")

    # AL: DSCR Yr1 = NOI Yr1 / DS Yr1
    dscr1_cell = _fc(ws, r, CAL,
        f"=IFERROR({fl(CAD)}/{fl(CAI)},0)", "0.00")
    dscr1_val = _safe_num(row_data.get("dscr_yr1"))
    if dscr1_val >= A.DSCR_MIN:
        dscr1_cell.fill = _GREEN
    elif dscr1_val >= A.DSCR_MIN - 0.10:
        dscr1_cell.fill = _YELLOW
    else:
        dscr1_cell.fill = _RED

    # AM: DSCR Yr3 = NOI Yr3 / DS Yr3 (IO year, same rate as Yr1)
    _fc(ws, r, CAM,
        f"=IFERROR({fl(CAF)}/{fl(CAI)},0)", "0.00")

    # AN: Exit Value = NOI Yr5 / exit_cap  (exit_cap embedded per deal)
    _fc(ws, r, CAN, f"=IFERROR({fl(CAH)}/{exit_cap:.6f},0)", "$#,##0")

    # AO: Loan Balance at Month 60 (formula: remaining balance after 24 amort payments)
    # Balance = PV of remaining 336 payments at same rate/payment
    _fc(ws, r, CAO,
        f"=IFERROR(-PV(Toggles!$B$16/12,360-24,PMT(Toggles!$B$16/12,360,-{fl(CY)})),0)",
        "$#,##0")

    # AP: Net Sale Proceeds = Exit * 99% - Loan Balance
    _fc(ws, r, CAP, f"={fl(CAN)}*0.99-{fl(CAO)}", "$#,##0")

    # AQ: Total Equity = W * 35%
    _fc(ws, r, CAQ, f"={fl(CW)}*0.35", "$#,##0")

    # AR: Deal IRR via XIRR
    # Cash flows: [-equity, NOI1-DS1, NOI2-DS1, NOI3-DS1, NOI4-DS4, NOI5-DS5+NSP]
    # Dates: annual from today
    xirr_vals = (
        f"CHOOSE({{1,2,3,4,5,6}},"
        f"-{fl(CAQ)},"
        f"{fl(CAD)}-{fl(CAI)},"
        f"{fl(CAE)}-{fl(CAI)},"
        f"{fl(CAF)}-{fl(CAI)},"
        f"{fl(CAG)}-{fl(CAJ)},"
        f"{fl(CAH)}-{fl(CAK)}+{fl(CAP)})"
    )
    xirr_dates = (
        "CHOOSE({1,2,3,4,5,6},"
        "DATE(ACQ_YEAR,ACQ_MONTH,ACQ_DAY),"
        "DATE(ACQ_YEAR+1,ACQ_MONTH,ACQ_DAY),"
        "DATE(ACQ_YEAR+2,ACQ_MONTH,ACQ_DAY),"
        "DATE(ACQ_YEAR+3,ACQ_MONTH,ACQ_DAY),"
        "DATE(ACQ_YEAR+4,ACQ_MONTH,ACQ_DAY),"
        "DATE(ACQ_YEAR+5,ACQ_MONTH,ACQ_DAY))"
    )
    _fc(ws, r, CAR,
        f"=IFERROR(XIRR({xirr_vals},{xirr_dates}),0)",
        "0.00%")

    # AS: IRR % (same value, explicit % format for display)
    _fc(ws, r, CAS, f"={fl(CAR)}", "0.00%")

    # AT: PASS / FAIL
    pf = _fc(ws, r, CAT, f'=IF({fl(CAR)}>=0.19,"PASS","FAIL")')
    irr_py = _safe_num(row_data.get("deal_irr"))
    pf.fill = _GREEN if irr_py >= A.DEAL_IRR_MIN else _RED

    # AU: Expense Ratio Yr1 = Fwd Exp / (NOI + Fwd Exp)
    exp_r = _fc(ws, r, CAU,
        f"=IFERROR({fl(CAC)}/({fl(CAD)}+{fl(CAC)}),0)", "0.00%")
    exp_py = _safe_num(row_data.get("expense_ratio_yr1"))
    if exp_py <= A.EXPENSE_RATIO_GREEN:
        exp_r.fill = _GREEN
    elif exp_py <= A.EXPENSE_RATIO_YELLOW:
        exp_r.fill = _YELLOW
    else:
        exp_r.fill = _RED

    # AV: Debt Yield Yr3 = NOI Yr3 / Loan Amount
    _fc(ws, r, CAV, f"=IFERROR({fl(CAF)}/{fl(CY)},0)", "0.00%")

    # AW: Going-In Cap (forward) = NOI Yr1 / Implied Value
    _fc(ws, r, CAW, f"=IFERROR({fl(CAD)}/{fl(CW)},0)", "0.00%")

    # AX: Loan Maturity Flag
    _fc(ws, r, CAX,
        f'=IFERROR(IF(DATEVALUE({fl(CK)})<TODAY()+365,"NEAR TERM",""),"")')
    # Apply orange to K column if near-term (Python side, based on stored date)
    loan_mat_str = str(row_data.get("loan_maturity_date") or "")
    if loan_mat_str and loan_mat_str not in ("--", "nan", ""):
        try:
            mat_dt = datetime.strptime(loan_mat_str[:10], "%Y-%m-%d")
            days_away = (mat_dt - datetime.now()).days
            if days_away < 365:
                ws.cell(row=r, column=CK).fill = _ORANGE
        except ValueError:
            pass

    # AY: Fail reason (FAIL tab only)
    if is_fail:
        c = ws.cell(row=r, column=CAY,
                    value=str(row_data.get("screen_reasons", "")))
        c.border = _THIN
        c.alignment = _LEFT


def _write_formula_tab(ws, df: pd.DataFrame, is_fail: bool = False):
    """Write header + all deal rows on a PASS or FAIL tab."""

    # Write header row
    all_headers = _HEADERS + ([_FAIL_EXTRA_HEADER] if is_fail else [])
    for col, label, width in all_headers:
        _hdr_cell(ws, 1, col, label, width)

    if df.empty:
        ws.cell(row=2, column=1, value="No deals in this category.")
        ws.freeze_panes = "B2"
        return

    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        _write_deal_row(ws, row_idx, row_data.to_dict(), is_fail=is_fail)

    # AutoFilter on header row
    last_col = CAY if is_fail else CAX
    ws.auto_filter.ref = f"A1:{_CL[last_col]}1"

    # Freeze row 1 and column A
    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# SHEET 4 -- EXCLUDED
# ---------------------------------------------------------------------------

def _write_excluded_tab(ws, df: pd.DataFrame):
    """Simple excluded/fail listing -- no formulas."""
    cols = [
        ("property_name",   "Property Name",   30),
        ("market",          "Market",          24),
        ("units",           "Units",            8),
        ("vintage",         "Vintage",          8),
        ("pipeline_status", "Status",          10),
        ("excluded_reason", "Excluded Reason", 50),
        ("screen_reasons",  "Screen Flags",    50),
    ]
    for ci, (_, label, width) in enumerate(cols, 1):
        _hdr_cell(ws, 1, ci, label, width)

    if df.empty:
        ws.cell(row=2, column=1, value="No excluded deals.")
        return

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, (key, _, _) in enumerate(cols, 1):
            val = row.get(key, "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _THIN
            c.alignment = _LEFT if ci in (1, 6, 7) else _CTR
            if key == "pipeline_status":
                c.fill = _YELLOW if val == "EXCLUDED" else _RED

    ws.auto_filter.ref = f"A1:{_CL[len(cols)]}1"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_reports(df: pd.DataFrame, output_dir: str,
                     meta: dict = None) -> tuple[str, str]:
    """
    Generate Excel (with toggle formulas) and plain-text reports.

    Args:
        df:         DataFrame from ingest.ingest_csv()
        output_dir: Path to data/outputs/
        meta:       Optional dict -- treasury_rate, loan_rate, treasury_source, run_ts

    Returns:
        (xlsx_path, txt_path)
    """
    os.makedirs(output_dir, exist_ok=True)
    date_str  = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_path = os.path.join(output_dir, f"deal_pipeline_{date_str}.xlsx")
    txt_path  = os.path.join(output_dir, f"deal_summary_{date_str}.txt")

    print(f"\n[Report] Generating Excel: {xlsx_path}")

    df_pass = df[df["pipeline_status"] == "PASS"].copy()
    df_fail = df[df["pipeline_status"] == "FAIL"].copy()
    df_excl = df[df["pipeline_status"] == "EXCLUDED"].copy()

    n_pass = len(df_pass)
    n_fail = int((df["pipeline_status"] == "FAIL").sum())
    n_excl = int((df["pipeline_status"] == "EXCLUDED").sum())
    n_poss = int((df["pipeline_status"] == "POSSIBLE").sum())

    if meta is None:
        meta = {}
    meta.setdefault("run_ts", datetime.now().strftime("%Y-%m-%d %H:%M"))
    meta.update({
        "total_deals": len(df),
        "n_pass": n_pass, "n_possible": n_poss,
        "n_fail": n_fail, "n_excl": n_excl,
    })

    # ---- Build workbook ----
    wb = Workbook()
    # Remove default sheet
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    ws_toggles = wb.create_sheet("Toggles",    0)
    ws_pass    = wb.create_sheet("PASS Deals",  1)
    ws_fail    = wb.create_sheet("FAIL Deals",  2)
    ws_excl    = wb.create_sheet("EXCLUDED",    3)

    # ---- Toggles sheet ----
    _write_toggles_tab(ws_toggles, meta)

    # ---- Named ranges (must be added after Toggles sheet exists) ----
    _add_named_ranges(wb)

    # ---- Deal tabs ----
    _write_formula_tab(ws_pass, df_pass, is_fail=False)
    _write_formula_tab(ws_fail, df_fail, is_fail=True)

    # ---- Excluded tab ----
    _write_excluded_tab(ws_excl, df_excl)

    wb.save(xlsx_path)
    print(f"[Report] Excel saved: {xlsx_path}")

    # ---- Plain-text summary ----
    lines = [
        "=" * 70,
        "PORTICO DEAL PIPELINE SUMMARY",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "=" * 70, "",
        f"  Total Deals Processed : {len(df)}",
        f"  PASS                  : {n_pass}",
        f"  POSSIBLE              : {n_poss}",
        f"  FAIL                  : {n_fail}",
        f"  EXCLUDED              : {n_excl}", "",
    ]
    if n_pass > 0:
        lines += ["-" * 70, "PASS DEALS", "-" * 70, ""]
        for _, row in df_pass.iterrows():
            bid = row.get("max_bid") or row.get("asking_price")
            bid_str = f"${bid:,.0f}" if bid and str(bid) != "nan" else "N/A"
            bpu = row.get("max_bid_per_unit")
            bpu_str = f"${bpu:,.0f}" if bpu and str(bpu) != "nan" else "N/A"
            lines.append(f"  {row.get('property_name', 'N/A')}")
            lines.append(f"    Market        : {row.get('market', '?')}")
            lines.append(f"    Units/Vintage : {row.get('units','?')} / {row.get('vintage','?')}")
            lines.append(f"    Max Bid       : {bid_str}  ({bpu_str}/unit)")
            for metric, label in [
                ("deal_irr",          "Deal IRR"),
                ("lp_irr",            "LP IRR"),
                ("going_in_cap",      "Going-In Cap"),
                ("dscr_yr1",          "DSCR Yr1"),
                ("expense_ratio_yr1", "Expense Ratio"),
                ("historical_noi",    "Hist NOI"),
            ]:
                val = row.get(metric)
                if val and str(val) != "nan":
                    if metric in ("deal_irr", "lp_irr", "going_in_cap", "expense_ratio_yr1"):
                        lines.append(f"    {label:<16}: {val:.2%}")
                    elif metric == "dscr_yr1":
                        lines.append(f"    {label:<16}: {val:.2f}x")
                    else:
                        lines.append(f"    {label:<16}: ${val:,.0f}")
            lines.append("")

    lines += ["=" * 70, "END OF REPORT", "=" * 70]
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[Report] Text summary saved: {txt_path}")

    return xlsx_path, txt_path
